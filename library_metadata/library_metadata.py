import csv
import functools
import os

import molmass
import openpyxl
import PySimpleGUI as sg

# Class of custom errors
class CustomError(Exception):
    ''' Base class for custom errors '''
    pass
class DesignFileError(CustomError):
    ''' Raised when inputfile is of the wrong filetype '''
    pass
class MassLynx_Metadata_Error(CustomError):
    ''' Raised when metadata string too long for MassLynx '''
    pass
class CloseAllWindows(CustomError):
    ''' Raised to alert to close all sg windows, if applicable
        Currently not used in the library_metadata script '''
    pass

def add_Virscidian_substance(x, substlist):
    ''' Add substance to Virscidian metadata substance list '''
    if substlist == reactants:
        groupname = 'SM'
    elif substlist == ISTD:
        groupname = 'ISTD'
    elif substlist == products:
        groupname = 'Product'
    elif substlist == sideproducts:
        groupname = 'Side-product'

    dosedict['Group'] = groupname
    
    dosedict['MF'] = mf
    dosedict['MW'] = mw

def metadataname(subst):
    ''' Replace Virscidian metadata-forbidden characters "'" and ":" in chemical names '''
    return subst['Name'].replace('\'', '*').replace(':', '~')

def makemetadatalist(metadata):
    ''' Used in createmetadata()
        Creates initial Virscidian metadata list for one table row '''
    i = 0
    metadatalist = []
    
    for x in range(5):
        if i == len(metadata)-1: # Stop at end of metadatalist
            break
        mstring = 'SI='
        for n, m in enumerate(metadata[i:], start = i):
            i = n
            if len(mstring + m) <= 253:
                mstring += f'{m}:'
            else:
                if len(m) > 251:
                    error_m_too_long = True
                    i += 1 # Exclude strings too long for the field
                break
            if i == len(metadata)-1: # Stop loop at end
                break
        if len(mstring) > 3:
            metadatalist.append(mstring)
        
    if i != len(metadata) - 1:
        error_list_too_long = True

    return metadatalist

def createmetadata(metadata):
    ''' Creates and refines Virscidian metadata list for one table row '''
    metadatalist = makemetadatalist(metadata)
    metadatalist += [''] * (5 - len(metadatalist)) # Ensure length 5

    return metadatalist

''' GUI functions '''
def getfile(message, **kwargs):
    # kwargs
    defaulttext = kwargs['defaulttext'] if 'defaulttext' in kwargs else ''
    title = kwargs['title'] if 'title' in kwargs else ''
    file_types = kwargs['file_types'] if 'file_types' in kwargs else (None, None)
    initial_folder = kwargs['initial_folder'] if 'initial_folder' in kwargs else ''
    spacersize = (kwargs['spacersize'], None) if 'spacersize' in kwargs else (32, None)
    no_titlebar = kwargs['no_titlebar'] if 'no_titlebar' in kwargs else True

    layout = [  [sg.Text(message.strip(), right_click_menu = right), sg.Text('', size = spacersize, right_click_menu = right)],
                [sg.InputText('_', key = 'fin'), sg.FileBrowse(tooltip = 'Experimental design file', initial_folder = initial_folder, file_types = file_types)],
                [sg.Button('OK', bind_return_key = True), sg.Text('', size = (43, None), right_click_menu = right)]
                ]

    window = sg.Window(title, layout, no_titlebar = no_titlebar, keep_on_top = True, return_keyboard_events = True, right_click_menu = right)
    window.Finalize()
    window['fin'].update(defaulttext)
    window.TKroot.focus_force()
    window.Element('fin').SetFocus()
    while True:
        try:
            event, values = window.Read()
            if event == 'OK':
                break
            if event in [sg.WIN_CLOSED, 'Exit']:
                window.close()
                raise CloseAllWindows
        except KeyboardInterrupt:
            window.close()
            raise CloseAllWindows
    window.close()
    fin = values['fin']
    
    return fin

def permissionerrorpopup(error, filename):
    message = f'Permission denied: {filename}'
    message = ('\n  ').join(message[i:i+64] for i in range(0, len(message), 64)) # Line breaks
    message += '\nPlease close the file to continue.'
    
    layout = [  [sg.Text(message)],
                [sg.Text('', size = (3, None)), sg.Button('OK', bind_return_key = True), sg.Text('', size = (30, None))]
                ]

    window = sg.Window(type(error).__name__, layout, keep_on_top = True, return_keyboard_events = True, right_click_menu = right)
    window.Finalize()
    window.TKroot.focus_force()

    while True:
        try:
            event, values = window.Read()
            if event in ['OK', sg.WIN_CLOSED, 'Escape:27', chr(13)]:
                break
            elif event == 'Exit':
                window.close()
                raise CloseAllWindows
        except (KeyboardInterrupt, CloseAllWindows):
            closeallwindows(window)
            return True # Error thrown
    window.close()

    return False # Error not thrown

''' Open output files '''
def openInExcel(func):
    ''' Decorator function to open files in Excel'''
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        value = func(*args, **kwargs)
        os.system(f'start excel "{value}"')
    return wrapper

@openInExcel
def openmasslynx():
    ''' Opens MassLynx file '''
    return os.path.join(anal_input_dir, f'{expt}.csv')

default_sms = [{'Name': 'mySM', 'Type': 'SM', 'MF': 'C10H12O'}]
default_ISTDs = []
default_pdts = [{'Name': 'myDP', 'Type': 'pdt', 'MF': 'C12H13Br'}]
default_imps = [{'Name': 'mySidePdt', 'Type': 'imp', 'MF': 'C12H14'}]

defaultsamplelist = [{'Location': 1, 'reactants': default_sms, 'ISTDs': default_ISTDs, 'pdts': default_pdts, 'imps': default_imps},
                     {'Location': 2, 'reactants': [{'Name': 'mySM', 'Type': 'SM', 'MF': 'C10H14O'}], 'ISTDs': default_ISTDs, 'pdts': default_pdts, 'imps': default_imps},
                     {'Location': 3, 'reactants': [{'Name': 'mySM', 'Type': 'SM', 'MF': 'C10H16O'}], 'ISTDs': default_ISTDs, 'pdts': default_pdts, 'imps': default_imps},
                     {'Location': 4, 'reactants': [{'Name': 'mySM', 'Type': 'SM', 'MF': 'C10H18O'}], 'ISTDs': default_ISTDs, 'pdts': default_pdts, 'imps': default_imps},
                     {'Location': 5, 'reactants': [{'Name': 'mySM', 'Type': 'SM', 'MF': 'C10H20O'}], 'ISTDs': default_ISTDs, 'pdts': default_pdts, 'imps': default_imps},
                     {'Location': 6, 'reactants': [{'Name': 'mySM', 'Type': 'SM', 'MF': 'C10H22O'}], 'ISTDs': default_ISTDs, 'pdts': default_pdts, 'imps': default_imps}
                     ]

samplelist = defaultsamplelist
expt = 'LIB-DEFAULT'
right = ['right', ['Continue', 'Exit']]
mydir = os.getcwd()

# Default MassLynx variables
method = 'CORTECS_T3_5' # Default method
MS_method = 'MS SCAN 1min' # Default
proc_method = 'RDM_Cortecs' # Default
plateloc = 3 # Default

# Load openpyxl workbook
# Set design file and directory
inputfile = None
myinputdir = None
newargs = {}
filetypes = ['.xlsm', '.xlsx']
while not inputfile:
    inputfilefound = True # Default value
    try:
        if not myinputdir:
            initial_folder = mydir
        else:
            initial_folder = myinputdir
        initial_folder = initial_folder if initial_folder.endswith(os.sep) else f'{initial_folder}{os.sep}'
        try:
            inputinputfile = getfile('Design file (xlsx/xlsm): ', 
                                        file_types = (('Excel files', [f'*{f}' for f in filetypes]),), 
                                        initial_folder = initial_folder, 
                                        spacersize = 32, 
                                        **newargs)
            inputinputfile = inputinputfile.replace('/', os.sep).replace('\\', os.sep)
        except CloseAllWindows:
            pass
        if not inputinputfile:
            raise DesignFileError
        inputinputfile = inputinputfile.replace('"', '') # Allow use of 'copy path' in Windows Explorer
        if f':{os.sep}' in inputinputfile:
            myinputdir = (os.sep).join(inputinputfile.split(os.sep)[:-1])
        elif os.sep * 2 in inputinputfile: # Server path
            myinputdir = (os.sep).join(inputinputfile.split(os.sep)[:-1])
        elif os.sep in inputinputfile:
            myinputdir = mydir + (os.sep).join(inputinputfile.split(os.sep)[:-1])
        inputfile = inputinputfile.split(os.sep)[-1]
        if myinputdir:
            mydir = myinputdir
            os.chdir(mydir)
        if '.' not in inputfile:
            inputfile = f'{inputfile}.xlsm' if os.path.isfile(f'{inputfile}.xlsm') else f'{inputfile}.xlsx'
        _, inputfiletype = os.path.splitext(inputfile)

        # Validate inputfile
        if not os.path.isfile(inputfile):
            newargs['defaulttext'] = inputinputfile
        if inputfiletype not in filetypes:
            filetypestring = joinListUserFriendly(filetypes)
            inputfiletype = ''
            newargs['defaulttext'] = inputinputfile
        if not os.path.isfile(inputfile) or not inputfiletype: # If file not found or wrong type
            newargs['defaulttext'] = inputinputfile
            inputfilefound = False
        if not inputfilefound:
            raise DesignFileError
    except DesignFileError:
        print(inputfile, inputinputfile)
        inputfile = None
    print(mydir)

# Load design workbook
while True:
    try:
        if inputfiletype == '.xlsm':
            wb = openpyxl.load_workbook(filename = inputfile, data_only = True, keep_vba = True)
        else:
            wb = openpyxl.load_workbook(filename = inputfile, data_only = True)
    except PermissionError as error:
        try: ### XXX
            permissionerrorpopup(error, inputfile)
        except CloseAllWindows:
            pass
    except OSError as error:
        print(f'Windows error: {error.filename}')
    except openpyxl.utils.exceptions.InvalidFileException as error: # File cannot be opened
        print(f'Openpyxl error: {error}')
    else:
        break

expt = inputfile.partition('.')[0]
st = wb.worksheets[0]
max_row = st.max_row
max_col = st.max_column
headers = []
formula_headers = []

# Define headers from Excel
for x in range(max_col):
    header = st.cell(row = 1, column = x + 1).value
    if header is not None:
        header = header.strip()
        if not header.isupper():
            header = header.capitalize()
        headers.append(header)
for col, header in enumerate(headers):
    if header.lower().endswith('formula'):
        name = header.lower().partition('_formula')[0]
        formula_headers.append((col, name))

reactantrows = [header[0] for header in formula_headers if header[1].startswith('reactant')]
ISTDrows = [header[0] for header in formula_headers if header[1].startswith('ISTD')]
productrows = [header[0] for header in formula_headers if header[1].startswith('product')]
sideproductrows = [header[0] for header in formula_headers if header[1].startswith('by-product')]

# Define sample list from input
samplelist = []
for x in range(1, max_row):
    reactants = []
    ISTDs = []
    products = []
    sideproducts = []
    for reactant in [h for h in formula_headers if h[1].startswith('reactant')]:
        myreactant = {'Name': reactant[1], 'Type': 'SM', 'MF': st.cell(row=x+1,column=reactant[0]+1).value}
        reactants.append(myreactant)
    for ISTD in [h for h in formula_headers if h[1].startswith('ISTD')]:
        myISTD = {'Name': ISTD[1], 'Type': 'ISTD', 'MF': st.cell(row=x+1,column=ISTD[0]+1).value}
        ISTDs.append(myISTD)
    for product in [h for h in formula_headers if h[1].startswith('product')]:
        myproduct = {'Name': product[1], 'Type': 'Product', 'MF': st.cell(row=x+1,column=product[0]+1).value}
        products.append(myproduct)
    for sideproduct in [h for h in formula_headers if h[1].startswith('by-product')]:
        sideproductname = sideproduct[1].replace('by-product', 'sideproduct')
        mysideproduct = {'Name': sideproductname, 'Type': 'Side-product', 'MF': st.cell(row=x+1,column=sideproduct[0]+1).value}
        sideproducts.append(mysideproduct)

    sample = {'Location': x, 'reactants': reactants, 'ISTDs': ISTDs, 'pdts': products, 'imps': sideproducts}
        
    samplelist.append(sample)

# Create metadata
for row, sample in enumerate(samplelist):
    reactantlist = sample['reactants']
    ISTDlist = sample['ISTDs']
    productlist = sample['pdts']
    sideproductlist = sample['imps']

    # Set SCM, SRT, SRTW to unknown and create superlist
    Virscidian_substances = []
    for substlist in [reactantlist, ISTDlist, productlist, sideproductlist]:
        for n, substance in enumerate(substlist):
            substance = {**substance, 'SCM': 'RTandMS', 'SRT': '', 'SRTW': ''} # Default
            substlist[n] = substance
        Virscidian_substances = [*Virscidian_substances, *substlist]

    # Create substance type list
    ST_list = ['SM' for r in reactantlist] + ['ISTD' for i in ISTDlist] + ['Product' for p in productlist] + ['Side-product' for s in sideproductlist]
    ST = ('_').join(ST_list) # Substance type

    # Create substance name list
    SN_list = [metadataname(subst) for subst in Virscidian_substances]
    SN = ('_').join(SN_list) # Substance name

    # Create molecular formula list
    F_list = [subst['MF'] for subst in Virscidian_substances]
    F = ('_').join([F if F else '' for F in F_list])

    # Create peak colour list
    SC = ST # Modified from substance type list ST
    peakcolours = [('SM', 'DarkRed'), ('ISTD', 'Purple'), ('Product', 'Yellow'), ('Side-product', 'Blue')]
    for pair in peakcolours:
        SC = SC.replace(pair[0], pair[1])

    # Create lists for substance confirmation
    SCM_list = [s['SCM'] for s in Virscidian_substances]
    SCM = ('_').join(SCM_list) # Substance confirmation mode
    SRT_list = [s['SRT'] for s in Virscidian_substances]
    SRT = ('_').join(SRT_list) # Substance retention time
    SRTW_list = [s['SRTW'] for s in Virscidian_substances]
    SRTW = ('_').join(SRTW_list) # Substance retention time width

    # Create selection and assignment lists
    PSC = SCM.replace('RTandMS', 'All').replace('RTOnly', 'Closest') # Peak selection criteria
    # `<x>` means any generic signal of type x
    SIS = SCM.replace('RTandMS', '<EIC>').replace('RTOnly', 'TWC') # Substance identity signal
    ''' SAM:: 0: All, 1: MSOnly, 2: MSOnlyAndOther, 3: Identity '''
    SAM = ('_').join(['2' for i in Virscidian_substances]) # Signal assignment mode

    ''' RTC:    0: Full time range, 
                1: Centroid-fixed window width, 
                2: Start time-End time,
                3: Centroid-% of base width, 
                4: Centroid-% of 10% height width, 
                5: Centroid-% of FWHH   '''
    RTC = SCM.replace('RTandMS', '0').replace('RTOnly', '1') # Retention time criteria

    # Sample metadata
    CID = f'{expt}_crude'
    SET = ''
    SUBSET = ''

    metadata = [f'CID!{CID}', f'L!{sample["Location"]}', 'PT!Plate', f'SET!{SET}', f'SS!{SUBSET}', 
                f'F!{F}', f'SN!{SN}', f'ST!{ST}', f'SC!{SC}', f'SCM!{SCM}', f'SAM!{SAM}', 
                f'PSC!{PSC}', f'SIS!{SIS}', f'RTC!{RTC}', f'SRT!{SRT}', f'SRTW!{SRTW}']
    sample['metadata'] = createmetadata(metadata)

    # Write MassLynx instructions
    masses = [pdt['MF'] if pdt else '' for pdt in productlist][:3]
    masses += [None] * (3 - len(masses)) # Ensure length 3. Values after 3rd not used
    sample['masses'] = masses

    samplelist[row] = sample

# Create MassLynx input file
g_drive_anal_input = r"G:\Chem\PERSONAL SHARING DIRECTORIES\Nessa Carson"
if os.path.isdir(g_drive_anal_input): # TODO XXXXXX
    anal_input_dir = g_drive_anal_input
else:
    drivemessage = 'G:\\Chem\\' if 'G:' in g_drive_anal_input else g_drive_anal_input
    message = f'Cannot connect to {drivemessage}. Creating MassLynx input file in directory: \n  {mydir}'
    print('\n' + message)
    sg.Popup(message, title=f'Cannot connect to {drivemessage}')
    anal_input_dir = mydir

while True:
    try:
        with open(os.path.join(anal_input_dir, f'{expt}_crude.csv'), 'w', newline = '', encoding = 'ANSI', errors = 'replace') as fout:
            writer = csv.writer(fout)
            line = ('Index', 'FILE_NAME', 'SAMPLE_LOCATION', 'INJ_VOL', 'PREP_INJ_VOL',
                    'INLET_FILE', 'MS_FILE', 'MASS_A', 'MASS_B', 'MASS_C', 'PROCESS_PARAMS', 'PROCESS',
                    'SPARE_1', 'SPARE_2', 'SPARE_3', 'SPARE_4', 'SPARE_5')
            writer.writerow(line)
            # Repeat '' for blank masses and metadata
            line = (1, f'{expt}-BLANK-OPT', f'0{plateloc}:01', 0, 1600, method, MS_method,) + \
                    ((''),) * 3 + (proc_method, 'AutoPurify',) + ((''),) * 5
            writer.writerow(line) # Blank injection
            kwargs_ANSI = {'encoding': 'ANSI', 'errors': 'replace'}
            for n, sample in enumerate(samplelist, start = 1):
                location = '%03d' % sample['Location']
                line = (n + 1, f'{expt}_{location}_1_crude-OPT', f'0{plateloc}:%02d' % n, 1, 1600, 
                        method, MS_method, *(sample['masses'][:3]), proc_method, 'AutoPurify', 
                        *sample['metadata'][:5])
                encodedline = []
                for segment in line:
                    if isinstance(segment, str): # Replace all ANSI errors. Unfortunately replaces 'β' with 'ß'
                        encodedsegment = (segment.encode(**kwargs_ANSI)).decode(**kwargs_ANSI)
                    else:
                        encodedsegment = segment
                    encodedline.append(encodedsegment)
                writer.writerow(encodedline)
        
    except PermissionError as error:
        filename = os.path.join(anal_input_dir, f'{expt}.csv')
        try:
            permissionerrorpopup(error, filename)
        except CloseAllWindows:
            pass
    else:
        break
